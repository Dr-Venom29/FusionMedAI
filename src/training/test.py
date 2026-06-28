import sys
import time
import json
import numpy as np
import pandas as pd
from pathlib import Path
from typing import Dict, Any, Tuple, Optional
import torch
import torch.nn as nn
from torch.utils.data import DataLoader

# Ensure project root is in sys.path
sys.path.append(str(Path(__file__).resolve().parents[2]))

import src.config as config
from src.utils.logger import setup_logger
from src.data.dataloader import create_dataloaders
from src.models.model_factory import load_model
from src.training.checkpoint import load_checkpoint
from src.training.metrics import calculate_metrics
from src.utils.visualization import (
    plot_training_history,
    plot_confusion_matrix,
    plot_roc_curves
)

def test_model(dry_run: bool = False) -> Dict[str, Any]:
    """
    Evaluates the best checkpoint of the model on the test dataset.
    Computes performance metrics, measures latency/FPS, exports predictions table,
    plots curves, and registers the run in global CSV and Excel tables.
    
    Args:
        dry_run: If True, evaluates only on a minimal subset of test data.
        
    Returns:
        Dict[str, Any]: Evaluation metrics dictionary.
    """
    # 1. Initialize logger
    logger = setup_logger(
        name="Tester",
        log_file=config.RUN_DIR / "test.log"
    )
    logger.info("Starting baseline testing pipeline...")
    
    # 2. Initialize test loader
    logger.info("Initializing DataLoaders...")
    _, _, test_loader = create_dataloaders(
        batch_size=config.BATCH_SIZE,
        num_workers=config.NUM_WORKERS,
        pin_memory=config.PIN_MEMORY
    )
    
    # 3. Initialize model and load best weights
    logger.info(f"Loading model '{config.MODEL_NAME}'...")
    model = load_model(
        name=config.MODEL_NAME,
        num_classes=config.NUM_CLASSES,
        pretrained=False  # No need for pretrained since we load checkpoint
    )
    
    best_ckpt_path = config.BEST_CHECKPOINT
    if not best_ckpt_path.exists():
        logger.warning(f"Best checkpoint not found at '{best_ckpt_path}'. Falling back to last checkpoint.")
        best_ckpt_path = config.LAST_CHECKPOINT
        
    if not best_ckpt_path.exists():
        raise FileNotFoundError(f"No checkpoint found at '{best_ckpt_path}' or '{config.LAST_CHECKPOINT}'. Please train a model first.")
        
    logger.info(f"Loading checkpoint: {best_ckpt_path}")
    checkpoint_data = load_checkpoint(
        checkpoint_path=best_ckpt_path,
        model=model,
        device=config.DEVICE
    )
    model = model.to(config.DEVICE)
    model.eval()
    
    # 4. Profile Model Parameters
    params_dict = model.get_num_parameters()
    logger.info(
        f"Model Profile - Total Parameters: {params_dict['total']:,} | "
        f"Trainable Parameters: {params_dict['trainable']:,} | "
        f"Model Size: {params_dict['size_mb']:.2f} MB"
    )
    
    # 5. Evaluate on Test Dataset
    all_targets = []
    all_preds = []
    all_probs = []
    image_ids = []
    
    # Extract original image IDs from test CSV to match predictions
    test_df = pd.read_csv(config.TEST_SPLIT_CSV)
    test_image_codes = test_df[config.ID_COLUMN].tolist()
    
    # Select test subset if dry_run
    if dry_run:
        logger.warning("Dry-run mode active. Evaluating on first 5 samples only.")
        test_loader_run = [next(iter(test_loader))]
        test_image_codes = test_image_codes[:len(test_loader_run[0][0])]
    else:
        test_loader_run = test_loader
        
    latency_list = []
    
    logger.info("Executing forward pass on test data...")
    with torch.no_grad():
        for i, (inputs, targets) in enumerate(test_loader_run):
            inputs = inputs.to(config.DEVICE, non_blocking=True)
            
            # Measure inference latency for this batch
            start_time = time.time()
            outputs = model(inputs)
            batch_latency = (time.time() - start_time) * 1000 # in ms
            
            # Store individual image latencies
            batch_size = inputs.size(0)
            latency_list.append(batch_latency / batch_size)
            
            probs = torch.softmax(outputs, dim=1)
            _, preds = torch.max(outputs, 1)
            
            all_targets.extend(targets.numpy())
            all_preds.extend(preds.cpu().numpy())
            all_probs.extend(probs.cpu().numpy())
            
    all_targets = np.array(all_targets)
    all_preds = np.array(all_preds)
    all_probs = np.array(all_probs)
    
    # 6. Calculate Latency Metrics
    avg_latency = float(np.mean(latency_list))
    fps = 1000.0 / avg_latency if avg_latency > 0 else 0.0
    logger.info(f"Latency Profile - Average: {avg_latency:.2f} ms/image | Throughput: {fps:.2f} FPS")
    
    # 7. Compute Metrics
    logger.info("Calculating performance metrics...")
    metrics = calculate_metrics(y_true=all_targets, y_pred=all_preds, y_probs=all_probs)
    metrics["avg_latency_ms"] = avg_latency
    metrics["fps"] = fps
    metrics["total_params"] = params_dict["total"]
    metrics["size_mb"] = params_dict["size_mb"]
    
    # 8. Export Predictions CSV
    logger.info("Exporting predictions table...")
    # Extract prediction confidence (probability of predicted class)
    confidences = [all_probs[idx, pred] for idx, pred in enumerate(all_preds)]
    
    pred_export_df = pd.DataFrame({
        "image": test_image_codes[:len(all_targets)],
        "label": all_targets,
        "prediction": all_preds,
        "confidence": confidences,
        # Save all class probabilities as strings
        "prob_class_0": all_probs[:, 0],
        "prob_class_1": all_probs[:, 1],
        "prob_class_2": all_probs[:, 2],
        "prob_class_3": all_probs[:, 3],
        "prob_class_4": all_probs[:, 4],
    })
    pred_export_df.to_csv(config.RUN_PREDICTIONS_CSV, index=False)
    logger.info(f"Saved prediction outputs to {config.RUN_PREDICTIONS_CSV}")
    
    # 9. Plot Curves
    logger.info("Generating evaluation curves...")
    # Load training history from JSON
    history = {}
    if config.RUN_HISTORY_JSON.exists():
        with open(config.RUN_HISTORY_JSON, "r", encoding="utf-8") as f:
            history = json.load(f)
            
    if history:
        plot_training_history(history=history, save_dir=config.CURVES_DIR)
        logger.info(f"Saved history curves to {config.CURVES_DIR}")
        
    plot_confusion_matrix(
        cm=metrics["confusion_matrix"],
        class_names=config.CLASS_NAMES,
        save_path=config.CONFUSION_MATRIX_DIR / "confusion_matrix.png"
    )
    logger.info(f"Saved confusion matrix plot to {config.CONFUSION_MATRIX_DIR}")
    
    plot_roc_curves(
        y_true=all_targets,
        y_probs=all_probs,
        class_names=config.CLASS_NAMES,
        save_path=config.ROC_DIR / "roc_curve.png"
    )
    logger.info(f"Saved ROC curve plot to {config.ROC_DIR}")
    
    # 10. Register Experiment Run in Tables
    _register_run_metrics(metrics)
    
    # 11. Generate Markdown Report
    _generate_markdown_report(metrics)
    
    return metrics

def _register_run_metrics(metrics: Dict[str, Any]) -> None:
    """Saves metrics in baseline_results.csv and appends a row to the global Excel sheet."""
    # Read/Append to CSV
    row_data = {
        "run_version": config.RUN_VERSION,
        "run_name": config.RUN_NAME,
        "model": config.MODEL_NAME,
        "accuracy": metrics["accuracy"],
        "balanced_accuracy": metrics["balanced_accuracy"],
        "precision": metrics["precision"],
        "recall": metrics["recall"],
        "f1_score": metrics["f1"],
        "qwk": metrics["qwk"],
        "auc": metrics.get("auc", 0.0),
        "sensitivity": metrics["sensitivity"],
        "specificity": metrics["specificity"],
        "avg_latency_ms": metrics["avg_latency_ms"],
        "fps": metrics["fps"],
        "total_params": metrics["total_params"],
        "model_size_mb": metrics["size_mb"],
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S")
    }
    
    # Write to local experiment baseline_results.csv
    local_baseline_csv = config.RUN_DIR / "baseline_results.csv"
    pd.DataFrame([row_data]).to_csv(local_baseline_csv, index=False)
    
    # Write/Append to global experiments/baseline_results.csv
    global_csv = config.BASELINE_CSV
    if global_csv.exists():
        global_df = pd.read_csv(global_csv)
        # Check if version already exists to update it, or append
        if config.RUN_VERSION in global_df["run_version"].values:
            global_df = global_df[global_df["run_version"] != config.RUN_VERSION]
        global_df = pd.concat([global_df, pd.DataFrame([row_data])], ignore_index=True)
    else:
        global_df = pd.DataFrame([row_data])
    global_df.to_csv(global_csv, index=False)
    
    # Write/Append to global Excel experiment sheet
    excel_path = config.EXPERIMENT_LOG_EXCEL
    if excel_path.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            sheet = wb.active
            
            # Find next free ID and append
            next_row = sheet.max_row + 1
            max_id = 0
            for r in range(2, sheet.max_row + 1):
                cell_val = sheet.cell(row=r, column=1).value
                if isinstance(cell_val, int):
                    max_id = max(max_id, cell_val)
                    
            next_id = max_id + 1
            
            # Date, Module, Experiment Name, Dataset, Model, Hyperparameters, Accuracy, Precision, Recall, F1 Score, AUC, ECE, Brier Score, Training Time, Notes, Status
            hyperparams_str = f"lr={config.LEARNING_RATE}, bs={config.BATCH_SIZE}, size={config.IMAGE_SIZE}"
            
            sheet.cell(row=next_row, column=1, value=next_id)
            sheet.cell(row=next_row, column=2, value=time.strftime("%Y-%m-%d"))
            sheet.cell(row=next_row, column=3, value="Retina")
            sheet.cell(row=next_row, column=4, value=config.RUN_NAME)
            sheet.cell(row=next_row, column=5, value="APTOS")
            sheet.cell(row=next_row, column=6, value=config.MODEL_NAME)
            sheet.cell(row=next_row, column=7, value=hyperparams_str)
            sheet.cell(row=next_row, column=8, value=metrics["accuracy"])
            sheet.cell(row=next_row, column=9, value=metrics["precision"])
            sheet.cell(row=next_row, column=10, value=metrics["recall"])
            sheet.cell(row=next_row, column=11, value=metrics["f1"])
            sheet.cell(row=next_row, column=12, value=metrics.get("auc", 0.0))
            sheet.cell(row=next_row, column=13, value="N/A")  # ECE
            sheet.cell(row=next_row, column=14, value="N/A")  # Brier
            sheet.cell(row=next_row, column=15, value="N/A")  # Training Time placeholder
            sheet.cell(row=next_row, column=16, value=f"Inference Latency: {metrics['avg_latency_ms']:.2f} ms")
            sheet.cell(row=next_row, column=17, value="Completed")
            
            wb.save(excel_path)
        except Exception as e:
            print(f"Error updating excel log sheet: {e}")

def _generate_markdown_report(metrics: Dict[str, Any]) -> None:
    """Generates the baseline_results.md markdown report under reports/."""
    report_path = config.REPORTS_DIR / "baseline_results.md"
    config.REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    
    cm = metrics["confusion_matrix"]
    cm_str = "\n".join([f"| Class {i} | " + " | ".join(map(str, row)) + " |" for i, row in enumerate(cm)])
    
    report_content = f"""# FusionMedAI: Baseline Model Evaluation Report

## Experiment Run Summary
* **Run Version**: `{config.RUN_VERSION}`
* **Run Description**: `{config.RUN_NAME}`
* **Model Backbone**: `{config.MODEL_NAME}`
* **Evaluation Timestamp**: `{time.strftime("%Y-%m-%d %H:%M:%S")}`
* **Target Device**: `{config.DEVICE}`

## Model Profile
* **Total Parameters**: {metrics['total_params']:,}
* **Trainable Parameters**: {metrics['total_params']:,}
* **Model Size**: {metrics['size_mb']:.2f} MB

## Overall Performance Metrics
| Metric | Value | Purpose / Description |
| :--- | :--- | :--- |
| **Quadratic Weighted Kappa (QWK)** | **{metrics['qwk']:.4f}** | Primary performance metric for ordinal classification. |
| **Macro F1-Score** | {metrics['f1']:.4f} | Balanced harmonic mean across all DR classes. |
| **Accuracy** | {metrics['accuracy']:.4f} | Overall ratio of correct predictions. |
| **Balanced Accuracy** | {metrics['balanced_accuracy']:.4f} | Accuracy adjusted for class imbalances. |
| **Sensitivity (Recall)** | {metrics['sensitivity']:.4f} | True positive rate averaged macro-wise. |
| **Specificity** | {metrics['specificity']:.4f} | True negative rate averaged macro-wise. |
| **Precision** | {metrics['precision']:.4f} | Positive predictive value averaged macro-wise. |
| **ROC AUC (Macro)** | {metrics.get('auc', 0.0):.4f} | Area under the Receiver Operating Characteristic curve. |

## Inference Speed / Latency Profile
* **Average Inference Speed**: `{metrics['avg_latency_ms']:.2f} ms/image`
* **Throughput**: `{metrics['fps']:.2f} FPS`

## Confusion Matrix (Counts)
| True \\ Predicted | No DR (0) | Mild (1) | Moderate (2) | Severe (3) | Proliferative (4) |
| :--- | :---: | :---: | :---: | :---: | :---: |
{cm_str}

## Visualization Files Saved
* Training History Curves: `experiments/{config.RUN_VERSION}_{config.RUN_NAME}/curves/`
* Confusion Matrix: `experiments/{config.RUN_VERSION}_{config.RUN_NAME}/confusion_matrix/confusion_matrix.png`
* ROC Curves: `experiments/{config.RUN_VERSION}_{config.RUN_NAME}/roc/roc_curve.png`
"""
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report_content)

if __name__ == "__main__":
    test_model(dry_run=True)
